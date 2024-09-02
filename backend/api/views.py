from openpyxl import load_workbook, Workbook
from collections import defaultdict

from django.http import HttpResponse
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser

from .models import User, Comercializadoras, Contrato, Factura, UVT
from .authentication import NitJWTAuthentication
from .serializers import *
from .filters import UserFilter, ContratoFilter, FacturaFilter


    
class UserView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filterset_class = UserFilter

class ComercializadorasView(generics.ListAPIView):
    queryset = Comercializadoras.objects.all()
    serializer_class = ComercializadorasSerializer
    
class CreateComercializadoraView(generics.CreateAPIView):
    queryset = Comercializadoras.objects.all()
    serializer_class = ComercializadorasSerializer
    
class UpdateComercializadoraView(generics.UpdateAPIView):
    queryset = Comercializadoras.objects.all()
    serializer_class = ComercializadorasSerializer

    
class DeleteComercializadoraView(generics.DestroyAPIView):
    queryset = Comercializadoras.objects.all()
    serializer_class = ComercializadorasSerializer

class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    permission_classes = (AllowAny,)
    serializer_class = RegisterSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
            user = serializer.save()
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            return Response({'error': e.detail}, status=status.HTTP_400_BAD_REQUEST)


class TokenView(TokenObtainPairView):
    serializer_class = TokenSerializer

        
class ContratoView(generics.ListAPIView):
    queryset = Contrato.objects.all()
    # permission_classes = [IsAuthenticated]
    serializer_class = ContratoSerializer
    filterset_class = ContratoFilter

    def post(self, request, *args, **kwargs):

        NITS = request.data.get("nits")

        if not NITS:
            NITS = ""

        contratos = Contrato.objects.filter(nit__nit__in=NITS).select_related('nit').prefetch_related('comercializadora')


        # consulta = '''
        #     SELECT * 
        #     FROM api_contrato 
        #     WHERE nit_id IN %s
        # '''

        consulta = '''
            SELECT 
                c.id, 
                c.nic,
                c.direccion as contrato_direccion,
                u.id as user_id,
                u.nit as user_nit,
                u.usuario as user_usuario,
                u.categoria as user_categoria,
                u.direccion as user_direccion,
                u.email as user_email
            FROM 
                api_contrato c
            INNER JOIN 
                api_user u
            ON 
                c.nit_id = u.nit
            WHERE 
                u.nit IN %s
        '''
        
        # Ejecutar la consulta con .raw()
        # contratos = Contrato.objects.raw(consulta, [tuple(NITS)])

        # Convertir RawQuerySet a lista de diccionarios
        contratosFinal = []
        for contrato in contratos:
            comercializadoras = contrato.comercializadora.all().values('id', 'nombre')
            contratosFinal.append({
                'contrato_id': contrato.id,  # Usar la clave primaria devuelta
                'nic': contrato.nic,
                'contrato_direccion': contrato.direccion,
                'user_id': contrato.nit.id,
                'user_nit': contrato.nit.nit,
                'user_usuario': contrato.nit.usuario,
                'user_categoria': contrato.nit.categoria,
                'user_direccion': contrato.nit.direccion,
                'user_email': contrato.nit.email,
                'comercializadora': list(comercializadoras)
            })

        return Response(contratosFinal)    


class FacturaView(generics.ListAPIView):
    queryset = Factura.objects.all()
    # permission_classes = [IsAuthenticated]
    serializer_class = FacturaSerializer
    filterset_class = FacturaFilter


    def post(self, request, *args, **kwargs):
        NICS = request.data.get("nics")

        facturas = Factura.objects.filter(nic__nic__in=NICS).select_related('nic__nit').prefetch_related('nic__comercializadora')

        facturas_list = []
        for factura in facturas:
            comercializadoras = factura.nic.comercializadora.all().values('id', 'nombre')
            facturas_list.append({
                'factura_id': factura.id,
                'anho': factura.anho,
                'data': factura.data,
                'contrato_id': factura.nic.id,
                'nic': factura.nic.nic,
                'contrato_direccion': factura.nic.direccion,
                'user_id': factura.nic.nit.id,
                'user_nit': factura.nic.nit.nit,
                'user_usuario': factura.nic.nit.usuario,
                'user_categoria': factura.nic.nit.categoria,
                'user_direccion': factura.nic.nit.direccion,
                'user_email': factura.nic.nit.email,
                'comercializadora': list(comercializadoras)
            })

        return Response(facturas_list)
    
        # consulta = '''
        #     SELECT *
        #     FROM 
        #         api_factura f
        #     INNER JOIN 
        #         api_user u
        #     INNER JOIN
        #         api_contrato c
        #     ON 
        #         c.nit_id = u.nit
        #     WHERE 
        #         nic_id IN %s
        # '''

        # facturas = Factura.objects.raw(consulta, [tuple(NICS)])

        # facturasFinal = []
        # for factura in facturas:
        #     facturasFinal.append({
        #         "data": factura.data
        #     })

        # print("FACTURA:", facturasFinal)
        # return Response(facturasFinal)


    
class UVTView(generics.ListAPIView):
    queryset = UVT.objects.all()
    serializer_class = UVTSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        anho_param = self.request.query_params.get('anho', None)

        if anho_param is not None:
            queryset = queryset.filter(anho=anho_param)

        return queryset
    


    
class TarifaView(APIView):
    
    def post(self, request, *args, **kwargs):

        KWH = float(request.data.get('KWH'))
        CATEGORIA = request.data.get('categoria')
        uvtPrecio = request.data.get('uvtPrecio')


        if CATEGORIA == "OFICIAL":
            return Response({
                "tarifa": 0,
                "inicio": 0,
                "fin": 0,
                "total": 0
            })

        TARIFAS = Tarifa.objects.filter(categoria=CATEGORIA)

        TARIFA_FINAL = []
        for tarifa in TARIFAS:
            if float(tarifa.inicio) <= KWH <= float(tarifa.fin):
                TARIFA_FINAL.append([tarifa.inicio, tarifa.fin, tarifa])
        

        TOTAL_A_PAGAR = int(float(str(TARIFA_FINAL[0][2])) * float(uvtPrecio))

        if TOTAL_A_PAGAR <= 100:
            TOTAL_A_PAGAR = round(TOTAL_A_PAGAR)
        elif 100 < TOTAL_A_PAGAR <= 10000:
            TOTAL_A_PAGAR = round(TOTAL_A_PAGAR / 100) * 100
        else:
            TOTAL_A_PAGAR = round(TOTAL_A_PAGAR / 1000) * 1000

        return Response({
            "tarifa": str(TARIFA_FINAL[0][2]),
            "inicio": TARIFA_FINAL[0][0],
            "fin": TARIFA_FINAL[0][1],
            "total": TOTAL_A_PAGAR}, status=status.HTTP_200_OK)
    

class uploadComercializadora(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):

        file = request.data.get('file')
        comercializadora = request.data.get('comercializadora')

        if not file:
            return Response({"error": "No se proporcionó ningún archivo."}, status=status.HTTP_400_BAD_REQUEST)


            
        try:
            wb = load_workbook(file)
            sheet = wb.active

            usuarios_data = {}
            contratos_data = {}
            facturas_data = defaultdict(lambda: defaultdict(list))

            for row in sheet.iter_rows(min_row=2, values_only=True):
                usuario, nit, direccion_pr, direccion_srv, nic, email, categoria, periodo, consumo, n1, n2, n3, n4 = row

                periodo_anho = str(periodo)[:4]
                periodo_mes = str(periodo)[4:]
                if periodo_mes.startswith("0"):
                    periodo_mes = str(periodo_mes[1:])
                
                categoria = categoria.upper()
                comercializadora = Comercializadoras.objects.get(nic=comercializadora)

                if nit not in usuarios_data:
                    usuarios_data[nit] = {
                        'usuario': usuario,
                        'direccion': direccion_pr,
                        'email': email,
                        'categoria': categoria,
                        'comercializadora': comercializadora,
                    }

                if nic not in contratos_data:
                    contratos_data[nic] = {
                        'direccion_srv': direccion_srv,
                        'nit': nit,
                        'comercializadora': comercializadora,
                    }

                facturas_data[nic][periodo_anho].append({
                    'mes': periodo_mes,
                    'consumo_KWH': consumo,
                    'categoria': categoria
                })


            usuario = ""

            for nit, user_data in usuarios_data.items():
                user, created = User.objects.update_or_create(
                    nit=nit,
                    defaults={
                        'usuario': user_data['usuario'],
                        'direccion': user_data['direccion'],
                        'email': user_data['email'],
                        'categoria': user_data['categoria'],
                    }
                )
                usuario = user_data
                user.comercializadora.set([user_data['comercializadora']])
                if created:
                    user.set_password("rootadmin")
                user.save()

            for nic, contrato_data in contratos_data.items():
                contrato, created = Contrato.objects.update_or_create(
                    nic=nic,
                    defaults={
                        'direccion': contrato_data['direccion_srv'],
                        'nit': User.objects.get(nit=contrato_data['nit']),
                    }
                )
                contrato.comercializadora.set([contrato_data['comercializadora']])
                contrato.save()

            for nic, facturas_por_anho in facturas_data.items():
                for anho, facturas in facturas_por_anho.items():

                    uvt = UVT.objects.get(anho=anho).valor

                    if facturas[0]["categoria"] == "OFICIAL":
                      
                        for factura in facturas:
                            factura["uvtPrecio"] = uvt
                            factura["inicio"] = 0
                            factura["fin"] = 0
                            factura["tarifa"] = 0
                            factura["totalPagar"] = 0

                    else:

                        TARIFAS = Tarifa.objects.filter(categoria=facturas[0]["categoria"])

                        TARIFA_FINAL = []
                        for factura in facturas:
                            for tarifa in TARIFAS:
                                if float(tarifa.inicio) <= factura["consumo_KWH"] <= float(tarifa.fin):
                                    TARIFA_FINAL.append([tarifa.inicio, tarifa.fin, tarifa.tarifa])
                        
                                    TOTAL_A_PAGAR = int(float(str(tarifa.tarifa)) * float(uvt))

                                    if TOTAL_A_PAGAR <= 100:
                                        TOTAL_A_PAGAR = round(TOTAL_A_PAGAR)
                                    elif 100 < TOTAL_A_PAGAR <= 10000:
                                        TOTAL_A_PAGAR = round(TOTAL_A_PAGAR / 100) * 100
                                    else:
                                        TOTAL_A_PAGAR = round(TOTAL_A_PAGAR / 1000) * 1000

                                    factura["uvtPrecio"] = uvt
                                    factura["inicio"] = tarifa.inicio
                                    factura["fin"] = tarifa.fin
                                    factura["tarifa"] = tarifa.tarifa
                                    factura["totalPagar"] = TOTAL_A_PAGAR

                    Factura.objects.update_or_create(
                        nic=Contrato.objects.get(nic=nic),
                        anho=anho,
                        data=facturas
                    )

        except Exception as e:
            return Response({"error": str(e)})
        
        finally:
            return Response({"success": "Datos almacenados en la base de datos éxitosamente."})
        



class ExportToExcelAPIView(APIView):
    def post(self, request, *args, **kwargs):
        data = request.data.get("data")

        wb = Workbook()
        ws = wb.active

        headers = ["NIT", "Usuario", "NIC", "Dirección", "Comercializadora", "Categoría", "Año", "Mes", 
                   "Consumo KWH", "UVT", "Tarifa UVT", "Rango", "Total a pagar"]
        
        ws.append(headers)

        for item in data:
            rango = f"{item['inicio']}-{item['fin']}"
            row = [
                item['nit'], item['usuario'], item['nic'], item['direccion'], item['comercializadora'],
                item['categoria'], item['anho'], item['mes'], item['consumo_KWH'], item['uvtPrecio'], 
                item['tarifa'], rango, item['totalPagar'],
            ]
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=datos_exportados.xlsx'
        
        wb.save(response)
        return response